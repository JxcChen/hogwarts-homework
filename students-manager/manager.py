# -*- coding:utf-8 -*-
# @Time     :2022/12/25 7:01 下午
# @Author   :CHNJX
# @File     :manager.py
# @Desc     :学员管理类
import os
import openpyxl

from student import Student


class Manager:
    __wb = None
    __ws = None
    __students_file_path = os.path.join(os.path.dirname(__file__), 'students.xlsx')

    def add_student(self, number: int, name: str, sex: str) -> str:
        """
        新增学员
        :param number: 编号
        :param name: 姓名
        :param sex: 性别
        :return: bool
        """
        if self.__check_student_correct(number, name, sex):
            raise Exception(self.__check_student_correct(number, name, sex))
        self.__get_students_file()
        # 判断学生是否已经存在
        if self.__is_student_exist(number) != -1:
            raise Exception('该学生已存在')
        cur_row = self.__ws.max_row + 1
        self.__ws.cell(cur_row, 1, number)
        self.__ws.cell(cur_row, 2, name)
        self.__ws.cell(cur_row, 3, sex)
        self.save_close_file()
        return '新增成功'

    def delete_student(self, number) -> str:
        """删除学生信息"""
        self.__get_students_file()
        res = self.__is_student_exist(number)
        if res == -1:
            return "该学生不存在"
        self.__ws.delete_rows(res)
        self.save_close_file()
        return '删除成功'

    def get_student(self, number):
        """获取学生信息"""
        self.__get_students_file()
        res = self.__is_student_exist(number)
        if res == -1:
            return "该学生不存在"
        row = self.__ws[res]
        stu = Student(row[0].value, row[1].value, row[2].value)
        self.close_file()
        return stu

    def get_student_list(self) ->list:
        """获取学生列表信息"""
        self.__get_students_file()
        student_list = []
        for row in self.__ws.iter_rows(min_row=2):
            stu = Student(row[0].value, row[1].value, row[2].value)
            print(stu)
            student_list.append(stu)
        self.close_file()
        return student_list

    def __get_students_file(self):
        self.__wb = openpyxl.load_workbook(self.__students_file_path)
        self.__ws = self.__wb.worksheets[0]

    def __is_student_exist(self, number) -> int:
        for i, row in enumerate(self.__ws.iter_rows(min_row=2)):
            if row[0].value == number:
                return i + 2
        return -1

    def __check_student_correct(self, number: int, name: str, sex: str) -> str:
        if type(number) is not int or number <= 0:
            return "学生编号输入错误"
        if type(name) is not str or name == '' or len(name) > 10:
            return "学生姓名输入错误"
        if sex != '男' and sex != '女':
            return "学生性别输入错误"
        return ''

    def save_close_file(self):
        self.__wb.save(self.__students_file_path)
        self.close_file()

    def close_file(self):
        self.__wb.close()

